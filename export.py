import csv
import io
from flask import send_file
from models import User, Subject, TestSession
from datetime import datetime


def export_by_students():
    subjects = Subject.query.order_by(Subject.order).all()
    sessions = (TestSession.query.filter_by(status='completed')
                .order_by(TestSession.completed_at.desc()).all())

    output = io.StringIO()
    output.write('\ufeff')  # BOM for Excel
    writer = csv.writer(output, delimiter=';')

    header = ['ФИО ученика', 'Дата рождения', 'Учитель', 'Дата тестирования']
    for subj in subjects:
        header += [f'{subj.name_ru} (балл)', f'{subj.name_ru} (макс)', f'{subj.name_ru} (%)']
    header += ['Итого', 'Макс. балл', 'Итого (%)']
    writer.writerow(header)

    for sess in sessions:
        row = [
            sess.student.full_name,
            sess.student.birth_date.strftime('%d.%m.%Y'),
            sess.teacher.full_name,
            sess.completed_at.strftime('%d.%m.%Y %H:%M') if sess.completed_at else ''
        ]
        subj_scores = {s.id: {'total': 0, 'count': 0} for s in subjects}
        for a in sess.answers:
            sid = a.question.subject_id
            if sid in subj_scores:
                subj_scores[sid]['total'] += a.score
                subj_scores[sid]['count'] += 1
        grand_total = 0
        grand_max = 0
        for subj in subjects:
            s = subj_scores[subj.id]
            s_max = s['count'] * 100
            s_pct = round(s['total'] / s_max * 100, 1) if s_max > 0 else 0
            row += [s['total'], s_max, s_pct]
            grand_total += s['total']
            grand_max += s_max
        grand_pct = round(grand_total / grand_max * 100, 1) if grand_max > 0 else 0
        row += [grand_total, grand_max, grand_pct]
        writer.writerow(row)

    fname = f'students_export_{datetime.now().strftime("%Y%m%d_%H%M")}.csv'
    
    bytes_output = io.BytesIO(output.getvalue().encode('utf-8'))
    bytes_output.seek(0)
    
    return send_file(
        bytes_output,
        as_attachment=True,
        download_name=fname,
        mimetype='text/csv'
    )


def export_by_teachers():
    subjects = Subject.query.order_by(Subject.order).all()
    teachers = User.query.filter_by(role='teacher').order_by(User.full_name).all()

    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output, delimiter=';')

    header = ['Учитель', 'Логин', 'Кол-во тестов']
    for subj in subjects:
        header += [f'{subj.name_ru} (ср. балл)']
    header += ['Ср. итог (%)']
    writer.writerow(header)

    for teacher in teachers:
        sessions = [s for s in teacher.sessions if s.status == 'completed']
        if not sessions:
            continue
        row = [teacher.full_name, teacher.username, len(sessions)]
        subj_totals = {s.id: {'total': 0, 'count': 0} for s in subjects}
        grand_pct_sum = 0
        for sess in sessions:
            for a in sess.answers:
                sid = a.question.subject_id
                if sid in subj_totals:
                    subj_totals[sid]['total'] += a.score
                    subj_totals[sid]['count'] += 1
            grand_pct_sum += sess.percentage()
        for subj in subjects:
            s = subj_totals[subj.id]
            avg = round(s['total'] / s['count'], 1) if s['count'] > 0 else 0
            row.append(avg)
        row.append(round(grand_pct_sum / len(sessions), 1) if sessions else 0)
        writer.writerow(row)

    fname = f'teachers_export_{datetime.now().strftime("%Y%m%d_%H%M")}.csv'
    
    bytes_output = io.BytesIO(output.getvalue().encode('utf-8'))
    bytes_output.seek(0)
    
    return send_file(
        bytes_output,
        as_attachment=True,
        download_name=fname,
        mimetype='text/csv'
    )


def export_print_report():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    
    subjects = Subject.query.order_by(Subject.order).all()
    sessions = (TestSession.query.filter_by(status='completed')
                .order_by(TestSession.completed_at.desc()).all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость"

    header = ['ФИО ученика', 'Учитель']
    for subj in subjects:
        header.append(subj.name_ru)
    header.append('Итого (%)')
    
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25

    for sess in sessions:
        row = [sess.student.full_name, sess.teacher.full_name]
        subj_scores = {s.id: {'total': 0, 'count': 0} for s in subjects}
        for a in sess.answers:
            sid = a.question.subject_id
            if sid in subj_scores:
                subj_scores[sid]['total'] += a.score
                subj_scores[sid]['count'] += 1
                
        for subj in subjects:
            s = subj_scores[subj.id]
            s_max = s['count'] * 100
            s_pct = round(s['total'] / s_max * 100, 1) if s_max > 0 else 0
            row.append(f"{s_pct}%")
            
        row.append(f"{sess.percentage()}%")
        ws.append(row)

        for col_idx in range(2, len(row) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.alignment = Alignment(horizontal='center')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    fname = f'printable_report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    
    return send_file(
        output,
        as_attachment=True,
        download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
